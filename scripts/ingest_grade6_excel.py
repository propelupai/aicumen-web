#!/usr/bin/env python3
"""
Ingest AIQ Grade 6 demo question bank (Excel) into Postgres.

Prerequisites:
  - migrations 005, 006, 007 applied
  - pip install -r scripts/requirements-ingest.txt
  - .env.local with DB_NAME, DB_USER, DB_PASS, DB_HOST (or Cloud SQL vars)

Usage:
  python scripts/ingest_grade6_excel.py --file "C:\\path\\AIQ_Demo_Question_Bank_Tagged.xlsx"
  python scripts/ingest_grade6_excel.py --file ... --status published
  python scripts/ingest_grade6_excel.py --file ... --dry-run
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as e:
    print("Missing dependency:", e, file=sys.stderr)
    print("Run: pip install -r scripts/requirements-ingest.txt", file=sys.stderr)
    sys.exit(1)

GRADE = 6
PACK_SLUG = "grade6_integrated_aiq"
BATCH_LABEL = "aiq_grade6_demo_excel_v1"
SHEET_NAME = "Demo Question Bank v3"

SUBJECT_TO_SLUG = {
    "English": "english",
    "Math": "maths",
    "Maths": "maths",
    "Science": "science",
    "Social Studies": "social_studies",
}

ACTIVITY_TYPE_MAP = {
    "Warm-Up": "warm_up",
    "CT Challenge": "ct_challenge",
    "Sports Spark": "sports_spark",
    "Debug It": "debug_it",
    "AI Connect": "ai_connect",
    "Exit Spark": "exit_spark",
    "Probe": "probe",
}

COLUMNS = {
    "external_id": "Question ID",
    "subject": "Subject",
    "chapter_code": "Chapter #",
    "chapter_title": "Chapter Title",
    "chapter_dependent": "Chapter-Dependent?",
    "activity_label": "Activity Type",
    "stem": "Question Text",
    "hint": "Hint / Nudge",
    "ct_skills": "CT Skill Tag(s)",
    "mandate_codes": "CBSE Mandate Code(s)",
    "mandate_labels": "CBSE Mandate Label(s)",
    "difficulty": "Difficulty",
    "socratic_branch": "Socratic Branch (Teacher-Only)",
    "answer": "Answer / Solution (Teacher-Only)",
    "source": "Source",
}


def load_env_local() -> None:
    root = Path(__file__).resolve().parents[1]
    env_path = root / ".env.local"
    if not env_path.exists():
        return
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, val = line.partition("=")
        key = key.strip()
        val = val.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = val


def connect():
    import psycopg2

    load_env_local()
    host = os.environ.get("DB_HOST")
    if host:
        return psycopg2.connect(
            host=host,
            dbname=os.environ["DB_NAME"],
            user=os.environ["DB_USER"],
            password=os.environ["DB_PASS"],
            sslmode="require",
        )
    raise SystemExit(
        "Set DB_HOST + DB_NAME + DB_USER + DB_PASS in .env.local for direct TCP ingest."
    )


def normalize_chapter_code(raw: str) -> str:
    t = (raw or "").strip()
    if re.match(r"^[\d.]+$", t):
        return t
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", t.lower())).strip("-")


def map_activity_type(label: str) -> tuple[str, str]:
    label = (label or "").strip()
    if label in ACTIVITY_TYPE_MAP:
        return ACTIVITY_TYPE_MAP[label], label
    if label.lower().startswith("puzzle:"):
        return "puzzle", label
    slug = re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", label.lower())).strip("_")
    return slug or "practice", label


def parse_ct_skills(raw: str | None) -> list[str]:
    if not raw:
        return []
    parts = re.split(r"[/,]", raw)
    out = []
    for p in parts:
        s = re.sub(r"_+", "_", re.sub(r"[^a-z0-9]+", "_", p.strip().lower())).strip("_")
        if s:
            out.append(s)
    return out


def parse_mandates(raw: str | None) -> list[str]:
    if not raw:
        return []
    return [c.strip().upper() for c in raw.split(",") if re.match(r"^M\d+$", c.strip(), re.I)]


def external_id_to_slug(external_id: str) -> str:
    return re.sub(r"-+", "-", re.sub(r"[^a-z0-9]+", "-", external_id.strip().lower())).strip("-")


def read_rows(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise SystemExit(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {name: headers.index(name) for name in COLUMNS.values()}

    out = []
    for i, row in enumerate(rows[1:], start=2):
        if not any(c is not None and str(c).strip() for c in row):
            continue
        rec = {key: row[col_idx[header]] for key, header in COLUMNS.items()}
        if not rec.get("external_id"):
            print(f"  skip row {i}: missing Question ID")
            continue
        out.append(rec)
    return out


def pg_json(value):
    from psycopg2.extras import Json
    return Json(value)


def upsert_batch(cur, source_file: str, row_count: int) -> int:
    cur.execute(
        """
        INSERT INTO content_ingest_batches (source_label, source_file, source_kind, row_count, notes)
        VALUES (%s, %s, 'excel', %s, %s)
        ON CONFLICT (source_label) DO UPDATE SET
          source_file = EXCLUDED.source_file,
          row_count   = EXCLUDED.row_count,
          ingested_at = NOW()
        RETURNING id
        """,
        (BATCH_LABEL, source_file, row_count, "AIQ Grade 6 demo bank from tagged Excel"),
    )
    return cur.fetchone()[0]


def get_subject_id(cur, excel_subject: str) -> int:
    slug = SUBJECT_TO_SLUG.get(excel_subject.strip())
    if not slug:
        raise ValueError(f"Unknown subject: {excel_subject}")
    cur.execute("SELECT id FROM subjects WHERE slug = %s", (slug,))
    row = cur.fetchone()
    if not row:
        raise ValueError(f"Subject slug '{slug}' not found — run migration 005")
    return row[0]


def upsert_chapter(cur, subject_id: int, chapter_code: str, chapter_title: str) -> int:
    cur.execute(
        """
        INSERT INTO chapters (subject_id, grade, chapter_code, title, metadata)
        VALUES (%s, %s, %s, %s, %s::jsonb)
        ON CONFLICT (subject_id, chapter_code) DO UPDATE SET
          title    = EXCLUDED.title,
          metadata = chapters.metadata || EXCLUDED.metadata,
          updated_at = NOW()
        RETURNING id
        """,
        (
            subject_id,
            GRADE,
            chapter_code,
            chapter_title.strip(),
            pg_json({"curriculum_pack": PACK_SLUG}),
        ),
    )
    return cur.fetchone()[0]


def build_answer_spec(answer: str | None) -> dict | None:
    if not answer or not str(answer).strip():
        return None
    text = str(answer).strip()
    return {"answer_type": "short_text", "body": {"value": text}}


def ingest_row(cur, rec: dict, batch_id: int, status: str, dry_run: bool) -> None:
    external_id = str(rec["external_id"]).strip()
    subject_id = get_subject_id(cur, str(rec["subject"]))
    chapter_code = normalize_chapter_code(str(rec["chapter_code"] or ""))
    chapter_title = str(rec["chapter_title"] or chapter_code).strip()
    chapter_dependent = str(rec.get("chapter_dependent") or "Yes").strip().lower() == "yes"

    activity_type, source_type_label = map_activity_type(str(rec.get("activity_label") or ""))
    slug = external_id_to_slug(external_id)
    ct_skills = parse_ct_skills(str(rec.get("ct_skills") or "") if rec.get("ct_skills") else None)
    mandates = parse_mandates(str(rec.get("mandate_codes") or "") if rec.get("mandate_codes") else None)
    mastery = str(rec.get("difficulty") or "").strip().lower() or None
    stem = str(rec.get("stem") or "").strip()
    hint = str(rec.get("hint") or "").strip() or None
    socratic = str(rec.get("socratic_branch") or "").strip() or None
    answer = str(rec.get("answer") or "").strip() or None
    source = str(rec.get("source") or "").strip() or None

    if not stem:
        raise ValueError(f"{external_id}: empty Question Text")

    title = f"{source_type_label}: {chapter_title}"[:200]
    metadata = {
        "quest_code": external_id,
        "theme": chapter_title,
        "difficulty": str(rec.get("difficulty") or "Developing").strip(),
        "stars": 2,
        "emoji": "📘",
        "accent": "teal",
        "source": source,
        "mandate_labels": str(rec.get("mandate_labels") or "").strip() or None,
    }
    ingest_raw = {
        "socratic_branch": socratic,
        "answer": answer,
        "source": source,
        "excel_activity_label": source_type_label,
        "mandate_labels": metadata.get("mandate_labels"),
    }

    if dry_run:
        print(f"  [dry-run] {external_id} → {slug} | {activity_type} | ch {chapter_code}")
        return

    chapter_id = upsert_chapter(cur, subject_id, chapter_code, chapter_title)

    cur.execute(
        """
        INSERT INTO activities (
          chapter_id, slug, title, activity_type, source_type_label,
          sort_order, estimated_minutes, ct_skills, status, metadata,
          external_id, chapter_dependent, mastery_level, ingest_batch_id,
          enrichment_status, ingest_raw
        ) VALUES (
          %s, %s, %s, %s, %s,
          %s, %s, %s, %s, %s::jsonb,
          %s, %s, %s, %s,
          'raw', %s::jsonb
        )
        ON CONFLICT (slug) DO UPDATE SET
          chapter_id         = EXCLUDED.chapter_id,
          title              = EXCLUDED.title,
          activity_type      = EXCLUDED.activity_type,
          source_type_label  = EXCLUDED.source_type_label,
          ct_skills          = EXCLUDED.ct_skills,
          status             = EXCLUDED.status,
          metadata           = EXCLUDED.metadata,
          external_id        = EXCLUDED.external_id,
          chapter_dependent  = EXCLUDED.chapter_dependent,
          mastery_level      = EXCLUDED.mastery_level,
          ingest_batch_id    = EXCLUDED.ingest_batch_id,
          enrichment_status  = CASE
            WHEN activities.enrichment_status = 'enriched' THEN activities.enrichment_status
            ELSE 'raw'
          END,
          ingest_raw         = EXCLUDED.ingest_raw,
          updated_at         = NOW()
        RETURNING id
        """,
        (
            chapter_id,
            slug,
            title,
            activity_type,
            source_type_label,
            int(re.search(r"-(\d+)$", external_id).group(1)) if re.search(r"-(\d+)$", external_id) else 0,
            10 if activity_type == "puzzle" else 15,
            ct_skills,
            status,
            pg_json(metadata),
            external_id,
            chapter_dependent,
            mastery,
            batch_id,
            pg_json(ingest_raw),
        ),
    )
    activity_id = cur.fetchone()[0]

    cur.execute("DELETE FROM questions WHERE activity_id = %s", (activity_id,))
    cur.execute(
        """
        INSERT INTO questions (activity_id, role, sort_order, stem, hint, answer_spec, context)
        VALUES (%s, 'stem', 0, %s, %s, %s::jsonb, %s::jsonb)
        RETURNING id
        """,
        (
            activity_id,
            stem,
            hint,
            pg_json(build_answer_spec(answer)) if answer else None,
            pg_json({"grade": GRADE, "chapter_dependent": chapter_dependent}),
        ),
    )

    if socratic:
        cur.execute(
            """
            INSERT INTO questions (
              activity_id, role, sort_order, label, stem, teacher_notes
            ) VALUES (%s, 'coach_step', 1, %s, %s, %s)
            """,
            (
                activity_id,
                "Socratic follow-up",
                "Use the branching prompts below based on the student's answer.",
                socratic,
            ),
        )

    cur.execute(
        "DELETE FROM activity_cbse_mandates WHERE activity_id = %s AND mandate_grade = %s",
        (activity_id, GRADE),
    )
    for code in mandates:
        cur.execute(
            """
            INSERT INTO activity_cbse_mandates (activity_id, mandate_grade, mandate_code)
            VALUES (%s, %s, %s)
            ON CONFLICT DO NOTHING
            """,
            (activity_id, GRADE, code),
        )

    print(f"  ok {external_id} (activity_id={activity_id}, mandates={','.join(mandates) or '-'})")



def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest AIQ Grade 6 Excel question bank")
    parser.add_argument("--file", required=True, help="Path to AIQ_Demo_Question_Bank_Tagged.xlsx")
    parser.add_argument(
        "--status",
        choices=("review", "published"),
        default="review",
        help="Initial activity status (default: review)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        raise SystemExit(f"File not found: {path}")

    rows = read_rows(path)
    print(f"Read {len(rows)} questions from {path.name}")

    if args.dry_run:
        for rec in rows:
            external_id = str(rec["external_id"]).strip()
            activity_type, _ = map_activity_type(str(rec.get("activity_label") or ""))
            chapter_code = normalize_chapter_code(str(rec["chapter_code"] or ""))
            print(f"  [dry-run] {external_id} -> {external_id_to_slug(external_id)} | {activity_type} | ch {chapter_code}")
        print("Dry run complete — no database changes.")
        return

    conn = connect()
    try:
        conn.autocommit = False
        with conn.cursor() as cur:
            batch_id = upsert_batch(cur, str(path), len(rows))
            print(f"Ingest batch id={batch_id} ({BATCH_LABEL})")

            for rec in rows:
                ingest_row(cur, rec, batch_id, args.status, dry_run=False)

        conn.commit()
        print(f"\nDone — {len(rows)} activities ingested (grade {GRADE}, status={args.status}).")
        print("Next: run enrichment worker on activities where enrichment_status = 'raw'.")
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
